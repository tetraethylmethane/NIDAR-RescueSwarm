"""Programme cost and unit economics, read from the BOM.

WHY THIS EXISTS
Business Strategy is worth 200 points -- as many as the entire Design Review --
and parameter 5, "Expenditure Breakdown & Resource Planning", is 20 of them and
explicitly requires a cost sheet. Everything here is arithmetic over the India
BOM, so a change to a supplier or a quantity moves the pitch numbers with it.

WHAT IS DERIVED AND WHAT IS NOT
Everything below comes from hardware/bom/RescueSwarm_BOM_India.xlsx and the tax
and contingency rates the BOM's own cost-summary tab states. Nothing here is a
market claim. Where the pitch needs one -- what a customer pays, how many units
a state buys, what a helicopter hour costs -- this module prints the question
instead of inventing an answer.

Run:  python tools/sizing-model/cost_model.py
"""
from __future__ import annotations

import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl needed:  python -m pip install --user openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
BOM = os.path.join(ROOT, "hardware", "bom", "RescueSwarm_BOM_India.xlsx")

N_AIRCRAFT = 3
# Rates as stated on the BOM's own '08 Cost Summary' tab.
DUTY_FREIGHT = 0.22      # on the IMPORTED residual only
GST = 0.18
CONTINGENCY = 0.20

# Consumables that recur per mission, from the BOM.
KIT_COUNT = 4            # kits carried per aircraft
W = 78


def rule(t=""):
    print("=" * W)
    if t:
        print(t)
        print("=" * W)


def tab_total(ws):
    """Sum qty x unit price, and the Indian-value share where declared."""
    qty_c = price_c = indig_c = None
    hdr = None
    for r in range(1, min(ws.max_row, 12) + 1):
        text = [str(ws.cell(r, c).value or "").replace("\n", " ").lower()
                for c in range(1, ws.max_column + 1)]
        for i, t in enumerate(text):
            if t.startswith("qty"):
                qty_c = i + 1
            if "unit price" in t or t == "unit cost (inr)":
                price_c = i + 1
            # Must be the PERCENTAGE column, not "Indigenisation gap or
            # action" further right, which is prose. Matching on "indig"
            # alone picked the last one and silently produced 0.0 % -- a
            # plausible answer, which is the dangerous kind of wrong.
            if indig_c is None and t.startswith("indig") and "%" in t:
                indig_c = i + 1
        if qty_c and price_c:
            hdr = r
            break
    if not (qty_c and price_c):
        return None
    total = indian = 0.0
    lines = 0
    for r in range(hdr + 1, ws.max_row + 1):
        q = ws.cell(r, qty_c).value
        p = ws.cell(r, price_c).value
        if isinstance(q, (int, float)) and isinstance(p, (int, float)):
            ext = q * p
            total += ext
            lines += 1
            if indig_c:
                f = ws.cell(r, indig_c).value
                if isinstance(f, (int, float)):
                    indian += ext * (f if f <= 1 else f / 100.0)
    return total, indian, lines


def inr(x):
    """Indian numbering is what the audience reads. Lakh = 1e5."""
    return f"INR {x:>11,.0f}  ({x / 1e5:>6.2f} L)"


if not os.path.exists(BOM):
    sys.exit(f"BOM not found: {BOM}")
wb = openpyxl.load_workbook(BOM, data_only=False)

rule("BOM ROLL-UP  -  Indian sourcing")
tabs = {}
for name in wb.sheetnames:
    got = tab_total(wb[name])
    if got:
        tabs[name] = got
        total, indian, lines = got
        print(f"  {name:<32}{lines:>3} lines  {inr(total)}")

air = tabs["01 Air Vehicle BOM"][0]
fleet_air = air * N_AIRCRAFT
one_off = sum(v[0] for k, v in tabs.items() if not k.startswith("01"))
subtotal = fleet_air + one_off

print()
print(f"  {'air vehicle, PER AIRCRAFT':<35}{inr(air)}")
print(f"  {'x' + str(N_AIRCRAFT) + ' for the fleet':<35}{inr(fleet_air)}")
print(f"  {'everything else, one-off':<35}{inr(one_off)}")
print(f"  {'SUBTOTAL':<35}{inr(subtotal)}")

rule("PROGRAMME COST  -  what the build actually needs funding for")
# Duty applies to the imported residual only, which is what indigenisation buys.
indian_value = sum(v[1] for v in tabs.values() if v[1])
declared = sum(v[0] for v in tabs.values() if v[1])
indig_frac = indian_value / declared if declared else 0.0
if indig_frac <= 0.0:
    sys.exit("indigenisation read as 0 % -- the BOM declares a fraction per "
             "line, so this is a parsing failure, not an answer. Check the "
             "'Indig. %' column detection before trusting any of the above.")
imported_residual = subtotal * (1 - indig_frac)
duty = imported_residual * DUTY_FREIGHT
gst = (subtotal + duty) * GST
conting = (subtotal + duty + gst) * CONTINGENCY
programme = subtotal + duty + gst + conting

print(f"  subtotal                          {inr(subtotal)}")
print(f"  indigenisation (value-weighted)   {indig_frac:>6.1%} of the priced BOM")
print(f"  imported residual                 {inr(imported_residual)}")
print(f"  duty + freight @ {DUTY_FREIGHT:.0%} on that      {inr(duty)}")
print(f"  GST @ {GST:.0%}                          {inr(gst)}")
print(f"  contingency @ {CONTINGENCY:.0%}                 {inr(conting)}")
print(f"  {'PROGRAMME TOTAL':<34}{inr(programme)}")
print()
print("  This is the funding ask. It is a competition build, not a product:")
print("  it carries test equipment, spares, a possible fourth aircraft and a")
print("  field data campaign that a production unit would not.")

rule("UNIT ECONOMICS  -  what an ADDITIONAL aircraft costs")
print("  The programme total is the wrong number to quote a customer, because")
print("  most of it is one-off. The marginal cost is what scales.")
print()
marginal = air * (1 + DUTY_FREIGHT * (1 - indig_frac)) * (1 + GST)
print(f"  marginal aircraft, taxes included   {inr(marginal)}")
print(f"  a 3-aircraft operational fleet      {inr(marginal * 3)}")
print(f"  + ground segment (one per fleet)    "
      f"{inr(tabs['03 Ground Segment'][0] * (1 + GST))}")
fleet_price_basis = marginal * 3 + tabs["03 Ground Segment"][0] * (1 + GST)
print(f"  {'DEPLOYABLE SYSTEM, cost basis':<35}{inr(fleet_price_basis)}")
print()
print("  Cost basis, not price. Margin, warranty, training, support and")
print("  certification are commercial decisions, not BOM arithmetic.")

rule("RECURRING COST  -  what a deployment consumes")
kits = tabs["02 Payload & Kits"][0]
print(f"  payload and kits, as bought         {inr(kits)}")
print()
print("  Per-mission consumption depends on three things this repo does not")
print("  know yet, and inventing them would be worse than leaving them open:")
print()
print("    * kit refill cost -- the BOM buys competition kits, not relief kits")
print("    * battery cycle life -- 21700 packs at this C-rate, in Indian heat")
print("    * crew and transport per deployment")
print()
print("  What IS known: the fleet carries 3 x 4 = 12 kits per sortie, hover")
print("  endurance is 15.3 min against a 7.7 min design mission, and the pack")
print("  is 292 Wh, so a sortie is a few rupees of electricity and a pack cycle.")
print("  The consumable that matters is kits, not energy.")

rule("INDIGENISATION  -  a scored differentiator, not a slogan")
print(f"  Value-weighted Indian content across the priced BOM: {indig_frac:.1%}")
print()
# What it is worth, rather than what it sounds like.
duty_full = subtotal * DUTY_FREIGHT
gst_full = (subtotal + duty_full) * GST
prog_full = (subtotal + duty_full + gst_full) * (1 + CONTINGENCY)
print(f"  {'programme as costed (61.3% Indian)':<38}{inr(programme)}")
print(f"  {'the same BOM fully imported':<38}{inr(prog_full)}")
print(f"  {'saved by sourcing in India':<38}{inr(prog_full - programme)}"
      f"   {(prog_full - programme) / prog_full:.0%}")
print()
print("  This is worth stating precisely because the audience is MeitY and the")
print("  Drone Federation of India, and because it is measurable rather than")
print("  asserted -- every line in the BOM carries a declared Indian fraction.")
print()
print("  It also buys schedule, which matters more than the money here: the")
print("  BOM's own note puts imports at 4-6 weeks against 2-3 weeks domestic,")
print("  inside a programme whose binding constraint is an 8-week flight window.")

rule("WHAT THIS MODULE WILL NOT DO")
print("  These are the pitch's remaining numbers, and they need real sources.")
print("  Each is a question for the team, not a gap to be filled with a guess:")
print()
print("    market size      how many flood-response units, at what refresh rate")
print("    willingness to   what NDRF/SDRF or a state disaster authority")
print("      pay            actually pays for comparable capability")
print("    alternative      cost per flight hour of the helicopter or boat")
print("      cost           search this displaces")
print("    adoption         DGCA category, BVLOS status, and what a state")
print("                     procurement cycle takes")
print()
print("  A pitch that derives its cost from a real BOM and admits which")
print("  numbers are unsourced is stronger than one that is confident")
print("  throughout, because the first can be checked.")
rule()
