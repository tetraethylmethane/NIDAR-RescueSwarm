#!/usr/bin/env python3
"""Stamp a SUPERSEDED banner across the legacy BOM workbooks.

These three describe configurations the project has moved off. They are kept
because they hold per-line detail the adopted budget summarises away -- named
parts, verified Indian suppliers and 28 working product links -- but the file
named "_Verified" is the most dangerous object in the directory: the name
invites trust and the number is 84 % above what is actually being asked for.

Anyone opening one should learn in the first cell that it is not the live
figure. Run again after any budget change; it re-reads the live ask.

    python hardware/bom/stamp_superseded.py
"""
from __future__ import annotations

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, os.path.join(ROOT, "docs", "proposal", "figures"))
import competition_budget as CB  # noqa: E402

SUB = sum(r[2] for _, rows in CB.GROUPS for r in rows)
SP = CB.tax_split()
DUTY = SP["excl"] * (1 - CB.INDIG) * CB.DUTY
GST = (SP["excl"] + DUTY) * CB.GST
ASK = round((SUB + DUTY + GST) * (1 + CB.CONTINGENCY))

LIVE = "RescueSwarm_Cost_Study.xlsx"
SRC = "docs/proposal/figures/competition_budget.py"

# file -> (per-aircraft figure it states, what it is still good for)
LEGACY = {
    "RescueSwarm_BOM.xlsx": (
        None,
        "The original pre-India costing. Superseded twice over."),
    "RescueSwarm_BOM_India.xlsx": (
        264_400,
        "First Indian-sourced pass. Superseded by the Verified sheet, "
        "then by the adopted configuration."),
    "RescueSwarm_BOM_India_Verified.xlsx": (
        290_546,
        "Fully specified variant with named parts, verified Indian suppliers "
        "and 28 product links. STILL THE BEST SOURCE for per-line sourcing "
        "detail -- but NOT for prices or totals."),
}

RED = "C00000"
YELLOW = "FFF2CC"


def stamp(path, stated, why):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    ws.insert_rows(1, 4)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    c = ws.cell(1, 1, "SUPERSEDED - DO NOT QUOTE THESE FIGURES")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=RED)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    delta = ""
    if stated:
        delta = (f" This sheet states {stated:,} per aircraft against the "
                 f"adopted {CB.PER_AIRCRAFT:,} -- "
                 f"{(stated/CB.PER_AIRCRAFT - 1) * 100:.0f} % high.")
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=10)
    c = ws.cell(2, 1,
                f"The live budget is {LIVE} (ask {ASK:,}, {ASK/1e5:.2f} L), "
                f"generated from {SRC}.{delta} {why}")
    c.font = Font(size=10)
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 18
    wb.save(path)
    return stated


if __name__ == "__main__":
    print(f"live ask: {ASK:,}  ({ASK/1e5:.2f} L)   "
          f"adopted per aircraft: {CB.PER_AIRCRAFT:,}\n")
    for name, (stated, why) in LEGACY.items():
        p = os.path.join(HERE, name)
        if not os.path.exists(p):
            print(f"  skip (absent)  {name}")
            continue
        try:
            wb = openpyxl.load_workbook(p)
            first = wb.worksheets[0].cell(1, 1).value
        except Exception as e:  # noqa: BLE001
            print(f"  FAILED to open {name}: {e}")
            continue
        if isinstance(first, str) and first.startswith("SUPERSEDED"):
            print(f"  already stamped  {name}")
            continue
        try:
            stamp(p, stated, why)
            print(f"  STAMPED          {name}"
                  + (f"   ({stated:,}/aircraft)" if stated else ""))
        except PermissionError:
            print(f"  LOCKED (open in Excel?)  {name}")
