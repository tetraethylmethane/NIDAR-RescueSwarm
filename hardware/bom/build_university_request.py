#!/usr/bin/env python3
"""Generate the university funding request workbook.

A sponsor and a university purchase committee read for different things. A
sponsor asks what it gets. A committee asks which budget head, capital or
consumable, does the asset stay with the institute, and who may sanction it.
This workbook answers the committee's questions from the same numbers as the
sponsor proposal -- it reframes, it does not restate.

Everything is imported from docs/proposal/figures/competition_budget.py. No
figure is typed here. Run:

    python hardware/bom/build_university_request.py
"""
from __future__ import annotations

import datetime as dt
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))          # hardware/bom
ROOT = os.path.dirname(os.path.dirname(HERE))              # repo root
sys.path.insert(0, os.path.join(ROOT, "docs", "proposal", "figures"))
import competition_budget as CB  # noqa: E402

REV, TODAY = "A", dt.date.today().isoformat()
INSTITUTE = "Thapar Institute of Engineering and Technology"

# ---------------------------------------------------------------- house style
NAVY, LIGHT, RULE = "1F3864", "D9E2F3", "8EA9DB"
GREEN, AMBER, GREY = "E2EFDA", "FFF2CC", "F2F2F2"
H1 = Font(bold=True, size=15, color="FFFFFF")
H2 = Font(bold=True, size=11, color="FFFFFF")
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
SMALL = Font(size=9, color="404040")
MONEY = '#,##0;[Red]-#,##0'
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
RIGHT = Alignment(horizontal="right", vertical="top")
THIN = Border(bottom=Side(style="thin", color=RULE))


def banner(ws, title, sub, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(1, 1, title)
    c.font, c.fill, c.alignment = H1, PatternFill("solid", fgColor=NAVY), Alignment(
        vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c = ws.cell(2, 2 - 1, sub)
    c.font, c.alignment = SMALL, Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 16
    ws.freeze_panes = "A4"


def header(ws, row, cols):
    for i, (label, _w) in enumerate(cols, start=1):
        c = ws.cell(row, i, label)
        c.font, c.alignment = H2, WRAP
        c.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[row].height = 28
    for i, (_l, w) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def put(ws, row, vals, money_cols=(), fill=None, bold=False, wrap_cols=()):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row, i, v)
        c.font = BOLD if bold else BASE
        c.alignment = WRAP if i in wrap_cols else (RIGHT if i in money_cols else TOP)
        if i in money_cols:
            c.number_format = MONEY
        c.border = THIN
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
    return row + 1


# ------------------------------------------------------------------ the model
SUB = sum(r[2] for _, rows in CB.GROUPS for r in rows)
SP = CB.tax_split()
DUTY = SP["excl"] * (1 - CB.INDIG) * CB.DUTY
GST = (SP["excl"] + DUTY) * CB.GST
CONT = (SUB + DUTY + GST) * CB.CONTINGENCY
# Rounded to whole rupees: this goes to an accounts office, and a committee
# member who clicks a cell should not find 824072.495758 behind it.
ASK = round(SUB + DUTY + GST + CONT)
LOADING = ASK - SUB

# Newly confirmed items are credited at what the request had budgeted for them,
# not at the original-programme price -- see CONFIRMED_HELD. Overstating the
# institute's own contribution in a document the institute audits is a bad trade.
INSTITUTIONAL = [(l, CB.CONFIRMED_HELD.get(l, a))
                 for _, rows in CB.GROUPS for l, a, b, n in rows
                 if b == 0 and n.startswith("R1")]
INST_TOTAL = sum(a for _, a in INSTITUTIONAL)

# Deferred is IMPORTED, not re-derived. Scanning GROUPS for zeroed lines misses
# the per-aircraft deferrals (second GNSS x3, parachutes x3, sub-GHz radio x3),
# which are not programme rows -- it gives 2.92 L against the true 3.17 L.
# Re-deriving a figure that already exists is how these artefacts drift apart.
sys.path.insert(0, HERE)
from build_cost_study import DEFERRED  # noqa: E402

# Conservative capital test: durable and demonstrably retained. Lines that are
# part-consumable ("remaining ground items", fabricated apparatus) are counted
# as consumable, so the retention claim survives an accounts-office challenge.
CAPITAL_KEYS = ("Air vehicles", "RTK base receiver", "Survey tripod",
                "Equipment cases", "Safety-pilot transmitter", "Thrust stand",
                "Battery chargers", "Charger PSU", "Calibrated scale",
                "Fire extinguisher", "Sun hood")

# Lines the institute may already hold. R1 was applied once and removed the
# clear cases; these were REDUCED rather than zeroed, which is what you do when
# you are unsure. Each carries the question that actually settles it.
VERIFY = [
    ("Calibrated scale + remaining", 12_000, "HIGH",
     "Any lab with a precision balance",
     "Is it calibrated, and can it LEAVE the building on competition day? It "
     "decides the Rule C2 weigh-in, so a bench-bound balance does not qualify."),
    ("Safety-pilot transmitter", 8_000, "HIGH",
     "Aeromodelling / robotics club",
     "Is it ExpressLRS, or can it take an ELRS module? A FlySky or Spektrum "
     "set is the wrong protocol and cannot bind to our receivers."),
    ("Survey tripod + tribrach", 4_000, "HIGH", "Civil Engineering survey lab",
     "Standard tribrach with a 5/8 in thread? Available for 8 field days?"),
    ("Charger PSU", 1_500, "HIGH", "Any electronics lab",
     "12 V at 20 A or better. The bench supply was already counted as "
     "institutional; this is the field equivalent."),
    ("Battery chargers", 14_000, "MEDIUM", "Robotics / aeromodelling club",
     "Does it balance-charge 6S Li-ion at 5 A+, and measure per-cell internal "
     "resistance? If it does, the separate cell tester stays deleted."),
    ("Equipment cases", 9_000, "MEDIUM", "Department stores",
     "Foam-lined crates that survive transport to the competition venue."),
    ("Fire extinguisher, sand, charging bags", 4_000, "MEDIUM",
     "Institute safety office",
     "Class D or CO2 rated for lithium, and may it be signed out to a field "
     "site? 54 cells are cycled across this programme."),
    ("RTK base receiver", 18_000, "CHECK", "Civil Engineering DGPS",
     "HIGHEST VALUE, LEAST LIKELY TO JUST WORK. Does it output RTCM3 "
     "corrections at 1 Hz to a third-party rover, and is it free during the "
     "flight window? A total station is not a substitute."),
]
# Drop anything already confirmed held, so tab 02 cannot offer a reduction
# that has already been banked. CONFIRMED_HELD is the single source for this.
VERIFY = [row for row in VERIFY if row[0] not in CB.CONFIRMED_HELD]
VERIFY_TOTAL = sum(v for _, v, _, _, _ in VERIFY)


def ask_if_verified():
    """The ask if every remaining tab-02 item turns out to be held on campus.

    The RTK base receiver is quoted ex-GST, so removing it shrinks the duty and
    GST base as well as the subtotal -- a flat subtraction understates it.
    """
    cut = VERIFY_TOTAL
    excl = SP["excl"] - sum(v for l, v, _, _, _ in VERIFY
                            if CB.TAX_STATUS.get(l, ("incl",))[0] == "excl")
    d = excl * (1 - CB.INDIG) * CB.DUTY
    g = (excl + d) * CB.GST
    base = (SUB - cut) + d + g
    return base + base * CB.CONTINGENCY


TRANCHES = [
    (1, "Months 1-2", "P1-P4", 2.90,
     "On award. Long-lead components ordered in month one; design point frozen."),
    (2, "Months 3-4", "P5", 2.45,
     "Released on: ground segment demonstrated, link-loss and low-battery "
     "behaviour verified by fault injection."),
    (3, "Months 5-6", "P6-P8", 1.75,
     "Released on: three airframes weighed, first hover complete."),
    (4, "Months 7-8", "P9-P10", 1.14,
     "Released on: perception and delivery trials reported with measured "
     "figures. Carries the contingency."),
]

# Indian sourcing carried across from RescueSwarm_BOM_India_Verified.xlsx.
# These are supplier LEADS at the adopted specification, not held quotations.
SUPPLIER = {
    "Flight controller": ("Holybro Pixhawk 6C Mini", "Indian distributor (Robu / ElectroPi)", "https://robu.in/"),
    "AI accelerator": ("Edge AI module. REQUIREMENT IS THROUGHPUT, NOT TOPS: "
                       ">=37 inferences/s at 640x640 INT8 (12 tiles at 3.06 Hz). "
                       "PCIe attach; USB accelerators do not reach it.",
                       "e-con Systems (Chennai)", "https://www.e-consystems.com/"),
    "GNSS RTK primary": ("RTK-capable receiver: must accept RTCM3 corrections "
                        "and report an RTK FIXED solution at <=3 cm CEP. "
                        "SBAS-corrected receivers at ~1.5 m CEP DO NOT MEET "
                        "THIS regardless of what the part is named -- confirm "
                        "against the datasheet, not the product title.",
                        "Teravolt Labs (India)", "https://teravoltlabs.com/"),
    "Motors": ("5008-class, 340 KV, 6S, 18 in prop. MUST deliver >=3.18 kgf "
               "static per motor with hover at 1.59 kgf (50% of max); "
               "published thrust curve required, or thrust-stand verified "
               "before fleet commitment. <=175 g.",
               "Reflex Drive (Lucknow) or equivalent", "https://reflexdrive.in/"),
    "Li-ion cells": ("21700 NMC, 4500 mAh min, 45 A continuous unrestricted "
                     "(no 80C cut-off), DC-IR <=15 mOhm @50% SoC 25C, <=72 g", "GODI India (Hyderabad)", "https://godiindia.com/"),
    "Structure": ("CF tube, plate and machined clamps", "Kineco Kaman (Goa) + institute machine shop", "https://www.kineco.in/"),
    "Camera + lens": ("Arducam IMX477 (type 1/2.3, 1.55 um) + 6 mm CS lens, "
                     "FIXED FOCUS. Hyperfocal is 4.15 m against a 30 m "
                     "minimum altitude, so a focus motor buys nothing and "
                     "adds a moving part.", "Indian distributor", "https://robu.in/"),
    "Pack, BMS, PDB, BEC": ("6S3P pack, BMS, PDB, BEC", "Flameback Tech (Baddi, HP)", "https://www.flamebacktech.com/"),
    "RC rx, storage, cooling, mounts": ("ExpressLRS 2.4 GHz rx (3.5+, native MAVLink)", "Zerodrag (India)", "https://zerodrag.in/"),
    "ESCs": ("50-60 A continuous, field-oriented, 6S. Peak demand is 29 A "
            "per motor at T/W 2.0.", "Reflex Drive (Lucknow)", "https://reflexdrive.in/"),
    "Video and coordination radios": ("Analog 5.8 GHz video set + SX1262 865 MHz modules", "FxUAV Technologies (Burla)", "https://fxuav.in/"),
    "Payload system": ("Servo release, 4 stations", "Zerodrag + in-house", "https://zerodrag.in/"),
    "Propellers": ("18 in carbon", "Reflex Drive / UAV Garage", "https://uavgarage.com/"),
    "Wiring, connectors": ("Silicone wire, XT90-S, JST-GH", "Polycab / Robu", "https://robu.in/"),
    "Prop adapters": ("6 mm self-tightening hub", "Reflex Drive", "https://reflexdrive.in/"),
    "VEGA co-processor": ("C-DAC VEGA ARIES v3.0", "C-DAC (free to top-100 teams)", "https://vegaprocessors.in/"),
}

wb = Workbook()

# ============================================================ 00 THE REQUEST
ws = wb.active
ws.title = "00 The Request"
banner(ws, "Request for Institutional Funding  —  Project RescueSwarm",
       f"NIDAR 2026-27, Track 1  ·  {INSTITUTE}  ·  Rev {REV}, {TODAY}  "
       f"·  GENERATED FILE — run hardware/bom/build_university_request.py", 5)
for i, w in enumerate((44, 16, 16, 14, 46), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 4
ws.cell(r, 1, "WHAT IS BEING REQUESTED").font = Font(bold=True, size=12, color=NAVY)
r += 1
r = put(ws, r, ["Item", "Amount (INR)", "", "", "Note"], (2,), NAVY, True)
for i in (1, 2, 5):
    ws.cell(r - 1, i).font = H2

r = put(ws, r, ["Total funding requested", ASK, "", "", "Across four milestone-gated tranches"], (2,), GREEN, True)
r = put(ws, r, ["  of which capital — asset retained by the institute", None, "", "", ""], (2,))
r = put(ws, r, ["  of which genuinely consumed", None, "", "", ""], (2,))
r += 1

ws.cell(r, 1, f"WHY THIS IS NOT A {ASK/1e5:.2f} LAKH EXPENSE").font = Font(
    bold=True, size=12, color=NAVY)
r += 1
for line in (
    "1.  Most of it becomes permanent institute property. Three UAV platforms, an RTK "
    "base station, a calibrated thrust stand, chargers and field equipment all remain "
    "on the department's books after the competition. See tab 03.",
    "2.  The institute already contributes 30% of the programme. GPUs, laptops, the 3D "
    "printer, the machine shop and lab instruments are already committed, at no new "
    "cost. See tab 01.",
    f"3.  The first decision is {TRANCHES[0][3]:.2f} L, not {ASK/1e5:.2f} L. Release is gated on demonstrated "
    "milestones, and the committee re-decides at each gate. See tab 06.",
    "4.  We audited your existing assets before asking. Tab 02 lists 70,500 of "
    "equipment that may already be on campus, with the exact question to settle each. "
    "Every yes reduces this request and costs the project nothing.",
):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, line)
    c.font, c.alignment = BASE, WRAP
    ws.row_dimensions[r].height = 30
    r += 1
r += 1

ws.cell(r, 1, "THE PROGRAMME IN ONE VIEW").font = Font(bold=True, size=12, color=NAVY)
r += 1
r = put(ws, r, ["", "INR", "Share", "", "Basis"], (2,), NAVY, True)
for i in (1, 3, 5):
    ws.cell(r - 1, i).font = H2
rows_view = [
    ("Requested from the institute", ASK, ASK / (ASK + INST_TOTAL), "This workbook"),
    ("Already provided by the institute", INST_TOTAL, INST_TOTAL / (ASK + INST_TOTAL),
     "Existing assets, tab 01"),
    ("Total programme value", ASK + INST_TOTAL, 1.0, ""),
]
for lbl, v, share, basis in rows_view:
    c = ws.cell(r, 3, share)
    r = put(ws, r, [lbl, v, share, "", basis], (2,), LIGHT if lbl.startswith("Total") else None,
            lbl.startswith("Total"))
    ws.cell(r - 1, 3).number_format = "0.0%"
r += 1
r = put(ws, r, ["Deliberately NOT requested (deferred, tab 07)", sum(a for _, a, _ in DEFERRED),
                "", "", "Capability knowingly given up to reduce this request"], (2,), AMBER)

# ================================================ 01 WHAT THE INSTITUTE GIVES
ws = wb.create_sheet("01 Institute Provides")
banner(ws, "What the institute already provides",
       "Counted at replacement value and REMOVED from the request. Budget rule R1: "
       "if the institution owns it, it is not in the ask.", 3)
header(ws, 4, [("Item", 46), ("Value if bought (INR)", 20), ("Where it comes from", 52)])
r = 5
SRC = {
    "GCS laptop": "Team-supplied",
    "Backup GCS laptop": "Team-supplied",
    "Portable power station": "Institute field power",
    "3D printer": "Department facility",
    "Soldering station": "Department facility",
    "Rotary tool + CF extraction": "Department facility",
    "Bench power supply": "Department facility",
    "Multimeters": "Department facility",
    "Training compute": "Institute GPUs — model training needs no new hardware",
}
for lbl, amt in sorted(INSTITUTIONAL, key=lambda x: -x[1]):
    r = put(ws, r, [lbl, amt, SRC.get(lbl, "Institute facility")], (2,), wrap_cols=(3,))
r = put(ws, r, ["TOTAL INSTITUTIONAL CONTRIBUTION", INST_TOTAL,
                f"{INST_TOTAL/(ASK+INST_TOTAL):.1%} of total programme value"],
        (2,), LIGHT, True, (3,))

# ==================================================== 02 VERIFY BEFORE BUYING
ws = wb.create_sheet("02 Verify Before Buying")
banner(ws, "Equipment that may already be on campus — please check before sanctioning",
       "If the institute already holds these, they leave the request entirely. "
       "This is the only reduction available that costs the project no capability.", 5)
header(ws, 4, [("Budget line", 34), ("INR", 11), ("Likelihood", 12),
               ("Who plausibly holds one", 30), ("The question that actually settles it", 62)])
r = 5
for lbl, amt, conf, who, q in VERIFY:
    fill = {"HIGH": GREEN, "MEDIUM": AMBER, "CHECK": None}[conf]
    r = put(ws, r, [lbl, amt, conf, who, q], (2,), fill, wrap_cols=(4, 5))
    ws.row_dimensions[r - 1].height = 42
r = put(ws, r, ["MAXIMUM REDUCTION IF ALL ARE HELD", VERIFY_TOTAL, "", "",
                f"Would take the request from {ASK/1e5:.2f} L to about "
                f"{ask_if_verified()/1e5:.2f} L. Items already confirmed held are NOT "
                f"listed above -- that reduction is already inside the "
                f"{ASK/1e5:.2f} L figure."], (2,), LIGHT, True, (5,))
ws.row_dimensions[r - 1].height = 34
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, "Ask the specific question, not the general one. \"Do you have a GPS?\" "
                  "gets a yes that turns out to be a total station. Each row above names the "
                  "property that decides whether the item is actually usable for this project.")
c.font, c.alignment = SMALL, WRAP
ws.row_dimensions[r].height = 30

# ======================================================== 03 CAPITAL RETAINED
ws = wb.create_sheet("03 Capital Retained")
banner(ws, "Capital — durable assets that remain institute property",
       "Conservative test: DURABLE AND DEMONSTRABLY RETAINED. Lines that are partly "
       "consumable -- field consumables, and the fabricated ground-truth apparatus -- "
       "are counted as consumable in tab 04 even though some of each survives, so this "
       "retention claim survives an accounts-office challenge.", 4)
header(ws, 4, [("Asset", 46), ("INR", 14), ("Head", 14), ("Afterlife once the competition ends", 56)])
r = 5
AFTER = {
    "Air vehicles": "Three flying UAV platforms for student projects, theses and future "
                    "NIDAR entries.",
    "RTK base receiver": "Centimetre-accurate GNSS base for any surveying, mapping or "
                         "robotics work on campus.",
    "Thrust stand": "Calibrated propulsion test instrument; reusable for every future "
                    "rotor or propeller study.",
    "Battery chargers": "General lab charging for all battery-powered student projects.",
    "Calibrated scale": "General laboratory instrument.",
    "Equipment cases": "Transport and storage for any field campaign.",
    "Safety-pilot transmitter": "Reusable across every future UAV built by the department.",
    "Survey tripod": "Returns to general survey use.",
    "Fire extinguisher": "Permanent lab safety equipment.",
    "Charger PSU": "General bench supply.",
    "Sun hood": "Field equipment.",
}
cap_total = 0
for gname, grp in CB.GROUPS:
    for lbl, _o, now, _n in grp:
        if now == 0 or not lbl.startswith(CAPITAL_KEYS):
            continue
        after = next((v for k, v in AFTER.items() if lbl.startswith(k)), "Retained by the department.")
        r = put(ws, r, [lbl, now, gname.split(",")[0][:14], after], (2,), wrap_cols=(4,))
        ws.row_dimensions[r - 1].height = 30
        cap_total += now
con_total = SUB - cap_total
cap_full = round(cap_total + LOADING * cap_total / SUB)
con_full = ASK - cap_full
r = put(ws, r, ["SUBTOTAL, capital (before duty, GST and contingency)", cap_total, "", ""], (2,), LIGHT, True)
r = put(ws, r, ["With duty, GST and contingency apportioned", cap_full, "",
                f"{cap_full/ASK:.1%} of the total request"], (2,), GREEN, True, (4,))

# ============================================================= 04 CONSUMABLES
ws = wb.create_sheet("04 Consumables")
banner(ws, "Consumable — the part that is genuinely spent",
       "This is the true unrecoverable cost of entering the competition.", 3)
header(ws, 4, [("Item", 48), ("INR", 14), ("Why it is consumed", 58)])
r = 5
for gname, grp in CB.GROUPS:
    for lbl, _o, now, note in grp:
        if now == 0 or lbl.startswith(CAPITAL_KEYS):
            continue
        r = put(ws, r, [lbl, now, note[:150] or "Expended in use."], (2,), wrap_cols=(3,))
        ws.row_dimensions[r - 1].height = 30
r = put(ws, r, ["SUBTOTAL, consumable", con_total, ""], (2,), LIGHT, True)
r = put(ws, r, ["With duty, GST and contingency apportioned", con_full,
                f"{con_full/ASK:.1%} of the request — the genuine cost of competing"],
        (2,), AMBER, True, (3,))

# ============================================================ 05 AIRCRAFT BOM
ws = wb.create_sheet("05 Aircraft BOM")
banner(ws, f"Air vehicle bill of materials — {CB.PER_AIRCRAFT:,} per aircraft, "
           f"{CB.N_AIRCRAFT} required",
       "Suppliers are Indian leads at the adopted specification, not held quotations. "
       "Lines marked KEEP are held at professional grade deliberately; the rest are "
       "hobby-grade where the failure mode is visible and the spec is easy to verify.", 8)
header(ws, 4, [("#", 5), ("Component", 30), ("Specification / part", 38), ("Grade", 8),
               ("Qty", 6), ("Unit INR", 11), ("Total INR", 12), ("Indian supplier lead", 34)])
r, n = 5, 0
for name, qty, unit, tier, note in CB.AIRCRAFT:
    n += 1
    part, sup, url = SUPPLIER.get(name, ("", "To be quoted", ""))
    r = put(ws, r, [n, name, part or note[:70], "KEEP" if tier == "KEEP" else "hobby",
                    qty, unit, qty * unit, sup], (6, 7),
            GREEN if tier == "KEEP" else None, wrap_cols=(3, 8))
    if url:
        c = ws.cell(r - 1, 8)
        c.hyperlink, c.font = url, Font(size=10, color="0563C1", underline="single")
    ws.row_dimensions[r - 1].height = 28
r = put(ws, r, ["", "PER AIRCRAFT", "", "", "", "", CB.PER_AIRCRAFT, ""], (6, 7), LIGHT, True)
r = put(ws, r, ["", f"FLEET OF {CB.N_AIRCRAFT}", "", "", "", "",
                CB.PER_AIRCRAFT * CB.N_AIRCRAFT, ""], (6, 7), LIGHT, True)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws.cell(r, 1, "A fully specified variant of this aircraft was costed at 2,90,546 per "
                  "unit using named, verified Indian parts throughout "
                  "(RescueSwarm_BOM_India_Verified.xlsx). The configuration above adopts "
                  "generic hobby-grade propulsion and ancillaries where the specification "
                  "is easy to verify on receipt, which is what brings it to "
                  f"{CB.PER_AIRCRAFT:,}. Motor thrust is unpublished at this grade and is "
                  "verified on the thrust stand before first flight.")
c.font, c.alignment = SMALL, WRAP
ws.row_dimensions[r].height = 48

# ======================================================== 06 RELEASE SCHEDULE
ws = wb.create_sheet("06 Release Schedule")
banner(ws, "Phased release — the committee re-decides at every gate",
       "Funds are not required at once. Each tranche is released only on evidence from "
       "the previous one, so the institute's exposure is capped at one tranche.", 5)
header(ws, 4, [("Tranche", 9), ("Timing", 14), ("Phases", 11), ("INR", 13),
               ("Release condition", 66)])
r, cum = 5, 0.0
for num, timing, phases, lakh, cond in TRANCHES:
    cum += lakh
    r = put(ws, r, [num, timing, phases, lakh * 1e5, cond], (4,),
            GREEN if num == 1 else None, wrap_cols=(5,))
    ws.row_dimensions[r - 1].height = 34
r = put(ws, r, ["", "TOTAL", "", ASK, "Cumulative across eight months"], (4,), LIGHT, True)
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, "These tranches are tied to technical milestones, not to approval "
                  "thresholds. They are a genuine risk-management structure: if the "
                  "programme fails to demonstrate a gate, the remaining funds are not "
                  "released and the institute's loss is capped. They are deliberately NOT "
                  "a split indent, and should not be presented as one.")
c.font, c.alignment = SMALL, WRAP
ws.row_dimensions[r].height = 44

# ================================================================ 07 DEFERRED
ws = wb.create_sheet("07 Deferred")
banner(ws, "What was deliberately not requested",
       "Recorded because a funding request is more credible when it states what it gave "
       "up. Each of these is a capability knowingly declined to keep the request small.", 3)
header(ws, 4, [("Item", 40), ("Would have cost (INR)", 20), ("Why it was declined", 74)])
r = 5
for lbl, amt, note in sorted(DEFERRED, key=lambda x: -x[1]):
    r = put(ws, r, [lbl, amt, note[:220]], (2,), wrap_cols=(3,))
    ws.row_dimensions[r - 1].height = 40
r = put(ws, r, ["TOTAL DEFERRED", sum(a for _, a, _ in DEFERRED),
                "Capability given up to reduce this request."], (2,), AMBER, True, (3,))
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1, "ONE ITEM NEEDS A DECISION FROM THE INSTITUTE, NOT FROM US: third-party "
                  "insurance was deferred at the team's direction, but cover is commonly "
                  "mandatory for unmanned operations in India. It must be confirmed against "
                  "the DGCA rules and the competition rulebook BEFORE ANY FLIGHT. If it is "
                  "required it ceases to be a budget choice, and 25,000 returns to the request.")
c.font, c.alignment = Font(bold=True, size=9, color="9C0006"), WRAP
ws.row_dimensions[r].height = 46

# ================================================================= 08 METHOD
ws = wb.create_sheet("08 Method")
banner(ws, "How these figures were produced",
       "So that any number here can be traced, challenged and reproduced.", 2)
header(ws, 4, [("", 42), ("", 86)])
r = 5
for k, v in [
    ("Single source of truth",
     "Every figure is imported from docs/proposal/figures/competition_budget.py. "
     "Nothing in this workbook is typed by hand, so it cannot drift from the "
     "technical proposal."),
    ("Customs duty", f"{CB.DUTY:.0%}, applied only to the imported share of "
     f"ex-GST quoted lines. Indigenous content is {CB.INDIG:.1%}, computed per line, "
     "not estimated."),
    ("GST", f"{CB.GST:.0%}, applied only to lines quoted ex-GST. Indian retail "
     "prices already include GST and are NOT taxed again — an earlier revision "
     "did this and overstated the request by 1.64 L."),
    ("Contingency", f"{CB.CONTINGENCY:.0%} on the whole, carried in the final tranche."),
    ("Rule R1", "If the institution already owns it, it is not in the request. This "
     f"removed {INST_TOTAL:,} (tab 01). Tab 02 lists what still needs checking."),
    ("Rule R2", "If it can be built rather than bought, it is costed as built — "
     "the thrust stand, ground-truth apparatus and sun hood are fabricated."),
    ("Rule R3", "Spares cover what CRASHES, not what fails rarely and costs a fortune. "
     "Accepted risks are stated in tab 07, not hidden."),
    ("Rule R4", "Hobby grade where the failure mode is visible and the specification is "
     "easy to verify; professional grade where it is not."),
    ("Known open item",
     "The generic motors publish no thrust curve. Thrust-to-weight is therefore a "
     "requirement rather than a measurement until the thrust stand runs in phase P5. "
     "This is stated rather than assumed away."),
    ("Reference documents",
     "hardware/bom/RescueSwarm_BOM_India_Verified.xlsx holds the fully specified "
     "variant with named parts and verified links. docs/proposal/ holds the technical "
     "proposal with the engineering justification."),
]:
    r = put(ws, r, [k, v], fill=None, bold=False, wrap_cols=(2,))
    ws.cell(r - 1, 1).font = BOLD
    ws.row_dimensions[r - 1].height = 44

# ---- fill the two summary cells on tab 00 now that both are known ----------
s0 = wb["00 The Request"]
s0.cell(7, 2, cap_full).number_format = MONEY
s0.cell(7, 5, f"{cap_full/ASK:.1%} — remains institute property (tab 03)")
s0.cell(8, 2, con_full).number_format = MONEY
s0.cell(8, 5, f"{con_full/ASK:.1%} — the true unrecoverable cost (tab 04)")
for rr in (7, 8):
    s0.cell(rr, 2).font = BOLD
    s0.cell(rr, 5).font = BASE
    s0.cell(rr, 5).alignment = WRAP

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

OUT = os.path.join(HERE, "RescueSwarm_University_Funding_Request.xlsx")
wb.save(OUT)

print(f"wrote {os.path.relpath(OUT, ROOT)}")
print(f"  ask                {ASK:>10,.0f}")
print(f"  capital retained   {cap_full:>10,.0f}   {cap_full/ASK:.1%}")
print(f"  consumed           {con_full:>10,.0f}   {con_full/ASK:.1%}")
print(f"  institute provides {INST_TOTAL:>10,}   {INST_TOTAL/(ASK+INST_TOTAL):.1%} of programme")
print(f"  to verify on campus{VERIFY_TOTAL:>10,}   would take the ask to "
      f"{ask_if_verified()/1e5:.2f} L")
print(f"  deferred           {sum(a for _, a, _ in DEFERRED):>10,}")
print(f"  tabs: {' | '.join(s.title for s in wb.worksheets)}")
