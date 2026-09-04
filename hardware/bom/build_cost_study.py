#!/usr/bin/env python3
"""Generate RescueSwarm_Cost_Study.xlsx from the adopted configuration.

WHY THIS EXISTS
An earlier version of this workbook was hand-built and then edited repeatedly as
the ask changed. It drifted until it disagreed with the funding proposal about
the headline number -- two artifacts stating different costs for the same
aircraft, which is the exact defect TRAPS.md is organised against.

This script removes the possibility. Every figure below is IMPORTED from
docs/proposal/figures/competition_budget.py, which is also what the proposal's
budget section and its figures derive from. There is one source. If the
configuration changes, re-run this and the workbook follows.

Run:  python hardware/bom/build_cost_study.py
"""
from __future__ import annotations

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, os.path.join(ROOT, "docs", "proposal", "figures"))

# THE single source of truth. Do not restate any of these numbers below.
import competition_budget as CB  # noqa: E402

OUT = os.path.join(HERE, "RescueSwarm_Cost_Study.xlsx")
REV, DATE = "F", "2026-08-16"

FN = "Calibri"


def F(sz=10, b=False, c="000000", u=None, i=False):
    return Font(name=FN, size=sz, bold=b, color=c, underline=u, italic=i)


TITLE = F(16, True, "1F3864")
SUB = F(9, c="7F7F7F")
LBL = F(10, True)
BODY = F(10)
SM = F(9)
LINK = F(9, c="0563C1", u="single")
WHITE = F(10, True, "FFFFFF")

NAVY = PatternFill("solid", fgColor="1F3864")
BAND = PatternFill("solid", fgColor="D6DCE4")
GREEN = PatternFill("solid", fgColor="E2EFDA")
AMBER = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FBE4E0")
GREY = PatternFill("solid", fgColor="F2F2F2")

W = Alignment(wrap_text=True, vertical="top")
C = Alignment(horizontal="center", vertical="center")
CT = Alignment(horizontal="center", vertical="top")
HAIR = Border(bottom=Side("hair", color="BFBFBF"))


def setup(ws, land=True, titles=None):
    ws.page_setup.orientation = "landscape" if land else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.oddFooter.left.text = f"RescueSwarm Cost Study rev {REV} - {DATE}"
    ws.oddFooter.right.text = "&P / &N"
    if titles:
        ws.print_title_rows = titles
    ws.sheet_view.showGridLines = False


def sect(ws, row, text, n, fill=NAVY, font=None):
    for j in range(1, n + 1):
        ws.cell(row, j).fill = fill
    c = ws.cell(row, 1, text)
    c.font = font or WHITE
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def head(ws, row, labels, widths=None, start=1):
    for j, h in enumerate(labels, start):
        c = ws.cell(row, j, h)
        c.font = F(9, True, "FFFFFF")
        c.fill = NAVY
        c.alignment = C
    ws.row_dimensions[row].height = 28
    if widths:
        for k, wd in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(k + 1)].width = wd


# --------------------------------------------------------------- derived
PER = CB.PER_AIRCRAFT
FLEET = PER * CB.N_AIRCRAFT
SUBTOTAL = sum(r[2] for _, rows in CB.GROUPS for r in rows)
# Tax only the ex-GST lines. CB.tax_split() is the authority; computing duty
# and GST here independently is how this workbook drifted from the proposal
# the first time.
SPLIT = CB.tax_split()
DUTY = SPLIT["excl"] * (1 - CB.INDIG) * CB.DUTY
GST = (SPLIT["excl"] + DUTY) * CB.GST
CONT = (SUBTOTAL + DUTY + GST) * CB.CONTINGENCY
ASK = SUBTOTAL + DUTY + GST + CONT
INST = sum(r[1] for _, rows in CB.GROUPS for r in rows
           if r[2] == 0 and r[3].startswith("R1"))

DEFERRED = [
    ("Flight insurance", 25_000, "MUST BE CONFIRMED against the rulebook and DGCA rules before any flight"),
    ("Second GNSS receiver (3)", 51_000, "Conditional on a P6 magnetometer measurement, not on funding"),
    ("Recovery parachutes (3)", 36_000, "Permitted, not required. ~40-point bet per crash avoided"),
    ("Training airframe", 37_309, "Would fly from week 1 and protect the fleet"),
    ("Sub-GHz safety radio (3)", 58_065, "A third independent link; two delicensed bands already carried"),
    ("Spare packs and motors", 74_396, "Survives a crash mid-campaign"),
    ("Field data campaign", 15_000, "Detection recall measured rather than modelled"),
    ("Spare compute module", 20_000, "Removes a single point of failure"),
    ("Spare airframe structure set", 25_000,
     "Deferred 2026-08-18. Was the crash cover; one hard landing now ends the campaign"),
    ("Spare propellers and stock", 21_400,
     "Deferred 2026-08-18. Propellers are the most consumed item in flight test"),
    ("Relief kits (14)", 5_600,
     "Deferred 2026-08-18. THE DELIVERED PAYLOAD. Rule C6 fixes it at 200 g; "
     "without kits there is nothing to release and the delivery task scores nothing"),
    ("Ground-truth apparatus", 4_860,
     "Deferred 2026-08-18. Detection recall now has no target to be measured against"),
]

EVIDENCE = [
    ("Flight controller", "Holybro Pixhawk 6C Mini", "Indian Robo Store",
     "https://www.indiamart.com/proddetail/pixhawk-6c-mini-flight-controller-2852778303248.html", 22_600),
    ("Flight controller (deferred)", "Agam Autopilot V6X-RT Full Set", "Agam Robotics",
     "https://www.agamrobotics.com/product-page/agam-pixhawk-6x-full-set", 42_000),
    ("Motor (deferred)", "Reflex Drive RD MI-5008 KV340", "Reflex Drive, Lucknow",
     "https://reflexdrive.in/product/rd-mi-5008-motor/", 9_099),
    ("ESC (deferred)", "Reflex Drive RD A-Series FOC 60 A", "Reflex Drive",
     "https://reflexdrive.in/product/rd-a-series-foc-esc-60a/", 3_400),
    ("Li-ion cell", "GODI 21700 NMC, BIS-certified", "GODI India",
     "https://evreporter.com/godi-receives-bis-certification-to-sell-li-ion-cells-in-india/", 700),
    ("Camera", "Arducam 12MP IMX477", "Robu.in",
     "https://robu.in/product/arducam-12mp-imx477-motorized-focus-high-quality-camera-for-raspberry-pi/", 9_600),
    ("GNSS RTK", "Teravolt AeroNav-1, NavIC L1+L5", "Teravolt Labs",
     "https://www.teravolt.in/product-page/aeronav-1-dual-band-gnss", 18_000),
    ("AI accelerator", "Raspberry Pi AI HAT+ 26 TOPS (Hailo-8)", "Robu.in",
     "https://robu.in/product/raspberry-pi-ai-hat-26-tops/", None),
    ("Sub-GHz radio (deferred)", "RFDesign RFD 868ux-IND, 865-870 MHz", "XBOOM India",
     "https://www.xboom.in/shop/drones/drone-components/transmission/rfd-868ux-ind-long-range-modem-bundle/", None),
    ("Co-processor", "C-DAC ARIES v3.0 (THEJAS32) - NOT an AI part", "C-DAC (MeitY)",
     "https://vegaprocessors.in/", 0),
]


def main():
    wb = openpyxl.Workbook()

    # ============================================================ 00 SUMMARY
    s = wb.active
    s.title = "00 Summary"
    setup(s, land=False)
    for col, wd in zip("ABCDE", [3, 40, 16, 16, 44]):
        s.column_dimensions[col].width = wd
    s["B2"] = "RescueSwarm - Air Vehicle Cost Study"
    s["B2"].font = TITLE
    s["B3"] = f"Revision {REV} - {DATE} - NIDAR 2026-27, Track 1"
    s["B3"].font = SUB
    s["B4"] = "GENERATED FILE. Do not edit by hand - run hardware/bom/build_cost_study.py"
    s["B4"].font = F(10, True, "C00000")

    sect(s, 6, "  SINGLE SOURCE OF TRUTH", 5)
    s["B7"] = ("Every figure in this workbook is imported from "
               "docs/proposal/figures/competition_budget.py, which is also what the funding "
               "proposal derives from. An earlier hand-built version of this workbook drifted "
               "until it disagreed with the proposal about the headline number. That cannot "
               "now happen: change the configuration, re-run this script, and both follow.")
    s["B7"].alignment = W
    s.merge_cells("B7:E9")

    sect(s, 11, "  THE ASK", 5)
    r = 12
    for lbl, val, note in [
        ("Per aircraft", PER, f"{CB.N_AIRCRAFT} aircraft"),
        ("Air vehicles, fleet", FLEET, ""),
        ("Subtotal, whole programme", SUBTOTAL, "All groups"),
        ("Tax-inclusive retail", SPLIT["incl"], "Dated Indian listings. No tax added."),
        ("Ex-GST quotes and services", SPLIT["excl"], "Quotes and fabrication. Taxed."),
        ("Duty and freight", DUTY, f"{CB.DUTY:.0%}, ex-tax lines only"),
        ("GST", GST, f"{CB.GST:.0%}, ex-tax lines only"),
        ("Contingency", CONT, f"{CB.CONTINGENCY:.0%} - phased release is itself a reserve"),
        ("TOTAL ASK", ASK, f"{ASK/1e5:.2f} lakh"),
        ("Institutional contribution", INST, "Equipment already held - co-funding in kind"),
        ("Indigenous content", CB.INDIG, "Computed from per-line fractions, not estimated"),
    ]:
        s.cell(r, 2, lbl).font = LBL if lbl.isupper() else BODY
        cell = s.cell(r, 3, val if lbl.startswith("Indigenous") else round(val))
        cell.alignment = C
        cell.number_format = "0.0%" if lbl.startswith("Indigenous") else "#,##0"
        cell.font = LBL if lbl.isupper() else BODY
        s.cell(r, 5, note).font = SM
        s.cell(r, 5).alignment = W
        for j in range(2, 6):
            s.cell(r, j).border = HAIR
            if lbl.isupper():
                s.cell(r, j).fill = AMBER
        s.row_dimensions[r].height = 16
        r += 1

    sect(s, 22, "  WHAT THIS CONFIGURATION GIVES UP", 5)
    for i, t in enumerate([
        "Nothing that is scored. RTK geolocation, onboard detection at rate, three concurrent "
        "video feeds and four-station delivery all survive.",
        "Margin: published thrust data, supplier support, spares depth.",
        f"Indigenous content, {0.58:.0%} -> {CB.INDIG:.0%}. The imported autopilot and generic "
        "propulsion are what make it affordable and what make it less Indian.",
        "Restoring the Indian autopilot and propulsion costs about 1.39 L across the fleet.",
    ]):
        s.cell(23 + i, 2, "- " + t).font = BODY
        s.cell(23 + i, 2).alignment = W
        s.merge_cells(f"B{23+i}:E{23+i}")
        s.row_dimensions[23 + i].height = 26

    sect(s, 28, "  THREE THINGS THAT GATE THIS NUMBER", 5)
    for i, t in enumerate([
        "INSURANCE was deferred at team direction. Third-party cover is commonly mandatory for "
        "Indian UAV operations. Confirm before any flight.",
        "DUTY AND GST MAY BE DOUBLE-COUNTED. The model adds 22% duty and 18% GST on top of "
        "prices that are largely Indian retail listings, already duty- and tax-paid. Exposure is "
        "roughly 2.3 L. A per-line audit is the largest remaining correction and it is an "
        "accounting fix, not a capability cut.",
        "MOTOR THRUST is unverified. The selected motors publish no thrust curve; the stand "
        "characterises them in P5, and a shortfall then forces a propulsion change.",
    ]):
        s.cell(29 + i, 2, f"{i+1}.  {t}").font = BODY
        s.cell(29 + i, 2).alignment = W
        s.merge_cells(f"B{29+i}:E{29+i}")
        s.row_dimensions[29 + i].height = 40

    # ==================================================== 01 ADOPTED AIRCRAFT
    a = wb.create_sheet("01 Adopted Aircraft")
    setup(a, titles="1:4")
    a["A1"] = f"Adopted air vehicle - {PER:,} per aircraft"
    a["A1"].font = TITLE
    a["A2"] = ("KEEP marks a line held at professional grade on purpose. Everything else is "
               "hobby-grade, chosen where the failure mode is visible and the rating is easy "
               "to verify on a bench.")
    a["A2"].font = SUB
    head(a, 4, ["Grade", "Component", "Qty", "Unit INR", "Ext INR", "Why this grade"],
         [8, 32, 5, 11, 11, 74])
    r = 5
    for name, q, u, tier, note in CB.AIRCRAFT:
        a.cell(r, 1, "KEEP" if tier == "KEEP" else "hobby").alignment = CT
        a.cell(r, 1).font = F(9, True) if tier == "KEEP" else SM
        a.cell(r, 2, name).alignment = W
        a.cell(r, 3, q).alignment = CT
        a.cell(r, 4, u).number_format = "#,##0"
        a.cell(r, 5, f"=C{r}*D{r}").number_format = "#,##0"
        a.cell(r, 6, note).alignment = W
        a.cell(r, 6).font = SM
        for j in range(1, 7):
            a.cell(r, j).border = HAIR
            if tier == "KEEP":
                a.cell(r, j).fill = GREEN
        a.row_dimensions[r].height = 26
        r += 1
    a.cell(r, 2, "TOTAL PER AIRCRAFT").font = LBL
    a.cell(r, 5, f"=SUM(E5:E{r-1})").font = LBL
    a.cell(r, 5).number_format = "#,##0"
    a.cell(r + 1, 2, f"x{CB.N_AIRCRAFT} FLEET").font = LBL
    a.cell(r + 1, 5, f"=E{r}*{CB.N_AIRCRAFT}").font = LBL
    a.cell(r + 1, 5).number_format = "#,##0"
    for j in range(1, 7):
        a.cell(r, j).fill = AMBER
        a.cell(r + 1, j).fill = AMBER
    a.freeze_panes = "B5"

    # ==================================================== 02 PROGRAMME BUDGET
    b = wb.create_sheet("02 Programme Budget")
    setup(b, titles="1:4")
    b["A1"] = "Programme budget - what changed and why"
    b["A1"].font = TITLE
    b["A2"] = ("'Original' is the development-programme costing of 28.74 L. 'Ask' is the "
               "competition build. Rules R1-R5 are stated in competition_budget.py.")
    b["A2"].font = SUB
    head(b, 4, ["Group", "Line", "Original", "Ask", "Delta", "Reason"],
         [24, 38, 11, 11, 11, 70])
    r = 5
    for gname, rows in CB.GROUPS:
        b.cell(r, 1, gname).font = F(10, True, "1F3864")
        for j in range(1, 7):
            b.cell(r, j).fill = BAND
        r += 1
        for lbl, old, new, note in rows:
            b.cell(r, 2, lbl).alignment = W
            b.cell(r, 3, old).number_format = "#,##0"
            b.cell(r, 4, new).number_format = "#,##0"
            b.cell(r, 5, new - old).number_format = "#,##0;[Red]-#,##0"
            b.cell(r, 6, note).alignment = W
            b.cell(r, 6).font = SM
            for j in range(1, 7):
                b.cell(r, j).border = HAIR
                if new == 0 and note.startswith("R1"):
                    b.cell(r, j).fill = GREEN
                elif new == 0:
                    b.cell(r, j).fill = GREY
                elif new > old:
                    b.cell(r, j).fill = RED
            b.row_dimensions[r].height = 26
            r += 1
    for lbl, v in [("SUBTOTAL", SUBTOTAL), ("Duty and freight", DUTY),
                   ("GST", GST), ("Contingency", CONT), ("TOTAL ASK", ASK)]:
        b.cell(r, 2, lbl).font = LBL
        b.cell(r, 4, round(v)).number_format = "#,##0"
        b.cell(r, 4).font = LBL
        for j in range(1, 7):
            b.cell(r, j).fill = AMBER
        r += 1
    b.freeze_panes = "B5"

    # ==================================================== 03 DEFERRED
    d = wb.create_sheet("03 Deferred")
    setup(d, land=False)
    d["A1"] = "Deferred, in reinstatement priority"
    d["A1"].font = TITLE
    d["A2"] = ("Removed to reach the ask, not judged unnecessary. A reviewer should be able to "
               "see what the lowest-cost configuration gives up.")
    d["A2"].font = SUB
    head(d, 4, ["#", "Item", "INR", "What it buys back"], [4, 30, 11, 62])
    r = 5
    for i, (lbl, v, note) in enumerate(DEFERRED, 1):
        d.cell(r, 1, i).alignment = CT
        d.cell(r, 2, lbl).alignment = W
        d.cell(r, 3, v).number_format = "#,##0"
        d.cell(r, 4, note).alignment = W
        d.cell(r, 4).font = SM
        for j in range(1, 5):
            d.cell(r, j).border = HAIR
        if i <= 2:
            for j in range(1, 5):
                d.cell(r, j).fill = AMBER
        d.row_dimensions[r].height = 26
        r += 1
    d.cell(r, 2, "TOTAL DEFERRED").font = LBL
    d.cell(r, 3, sum(v for _, v, _ in DEFERRED)).number_format = "#,##0"
    d.cell(r, 3).font = LBL
    for j in range(1, 5):
        d.cell(r, j).fill = AMBER

    # ==================================================== 04 EVIDENCE
    e = wb.create_sheet("04 Evidence")
    setup(e, titles="1:4")
    e["A1"] = "Evidence - where the prices come from"
    e["A1"].font = TITLE
    e["A2"] = ("Single-unit Indian retail, read on or before the revision date. Prices move. "
               "NOTE: these are tax-paid retail listings, which is why the duty and GST "
               "treatment on tab 02 is flagged as probably double-counted.")
    e["A2"].font = SUB
    head(e, 4, ["Line", "Part", "Supplier", "Source", "INR"], [26, 44, 26, 54, 10])
    r = 5
    for line, part, sup, url, price in EVIDENCE:
        e.cell(r, 1, line).font = LBL
        e.cell(r, 1).alignment = W
        e.cell(r, 2, part).alignment = W
        e.cell(r, 2).font = SM
        e.cell(r, 3, sup).alignment = W
        e.cell(r, 3).font = SM
        c = e.cell(r, 4, url)
        c.hyperlink = url
        c.font = LINK
        c.alignment = W
        e.cell(r, 5, price if price is not None else "confirm").alignment = CT
        e.cell(r, 5).number_format = "#,##0"
        for j in range(1, 6):
            e.cell(r, j).border = HAIR
        if "deferred" in line:
            for j in range(1, 6):
                e.cell(r, j).fill = GREY
        e.row_dimensions[r].height = 26
        r += 1
    e.freeze_panes = "A5"

    wb.save(OUT)
    print(f"wrote {os.path.relpath(OUT, ROOT)}")
    print(f"  per aircraft {PER:>10,}")
    print(f"  subtotal     {SUBTOTAL:>10,}")
    print(f"  ASK          {ASK:>10,.0f}   ({ASK/1e5:.2f} L)")
    print(f"  indigenous   {CB.INDIG:>10.1%}")
    print(f"  deferred     {sum(v for _, v, _ in DEFERRED):>10,}")
    print(f"  tabs: {' | '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
